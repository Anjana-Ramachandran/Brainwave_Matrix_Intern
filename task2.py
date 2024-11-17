import tkinter as tk
from tkinter import ttk, messagebox
from openpyxl import Workbook, load_workbook
import os

class Product:
    def __init__(self, name, quantity, price):
        self.name = name
        self.quantity = quantity
        self.price = price

class InventoryManager:
    def __init__(self, file_name='inventory.xlsx'):
        self.inventory = {}
        self.file_name = file_name
        self.load_inventory()

    def add_product(self, name, quantity, price):
        if name in self.inventory:
            messagebox.showwarning("Duplicate Entry", "Product already exists.")
        else:
            self.inventory[name] = Product(name, int(quantity), float(price))
            self.save_inventory()
            messagebox.showinfo("Success", f"Product '{name}' has been added.")

    def edit_product(self, name, quantity, price):
        if name in self.inventory:
            self.inventory[name].quantity = int(quantity)
            self.inventory[name].price = float(price)
            self.save_inventory()
            messagebox.showinfo("Success", f"Product '{name}' has been updated.")
        else:
            messagebox.showerror("Not Found", "Product not found.")

    def delete_product(self, name):
        if name in self.inventory:
            del self.inventory[name]
            self.save_inventory()
            messagebox.showinfo("Success", f"Product '{name}' has been deleted.")
        else:
            messagebox.showerror("Not Found", "Product not found.")

    def low_stock_alert(self, threshold=5):
        low_stock_products = [name for name, product in self.inventory.items() if product.quantity <= threshold]
        if low_stock_products:
            messagebox.showwarning("Low Stock Alert", f"Low stock products: {', '.join(low_stock_products)}")
        else:
            messagebox.showinfo("Stock Status", "All products are sufficiently stocked.")

    def save_inventory(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Inventory"
        ws.append(['Product Name', 'Quantity', 'Price'])

        for product in self.inventory.values():
            ws.append([product.name, product.quantity, product.price])

        wb.save(self.file_name)

    def load_inventory(self):
        if os.path.exists(self.file_name):
            wb = load_workbook(self.file_name)
            ws = wb.active

            for row in ws.iter_rows(min_row=2, values_only=True):
                name, quantity, price = row
                self.inventory[name] = Product(name, int(quantity), float(price))

class InventoryApp:
    def __init__(self, root):
        self.manager = InventoryManager()
        self.root = root
        self.root.title("Inventory Management System")
        self.root.configure(bg="ivory3")
        self.create_widgets()

    def create_widgets(self):
        # Product Form
        self.create_form()
        self.create_buttons()
        self.create_treeview()

    def create_form(self):
        labels = ["Product Name:", "Quantity:", "(â‚¹) Price:"]
        self.entries = []

        for idx, text in enumerate(labels):
            label = tk.Label(self.root, text=text, font=("Helvetica", 12), bg="ivory3")
            label.grid(row=idx, column=0, pady=10, padx=10, sticky='w')

            entry = tk.Entry(self.root, font=("Arial", 12))
            entry.grid(row=idx, column=1, pady=10, padx=10)
            self.entries.append(entry)

    def create_buttons(self):
        buttons = [
            ("Add Product", self.add_product, "green"),
            ("Edit Product", self.edit_product, "yellow"),
            ("Delete Product", self.delete_product, "orange red"),
            ("Low Stock Alert", self.low_stock_alert, "light coral"),
            ("Preview Inventory", self.preview_inventory, "cyan"),
            ("Exit", self.root.quit, "red")
        ]

        for idx, (text, command, color) in enumerate(buttons):
            btn = tk.Button(self.root, text=text, command=command, bg=color, fg="black" if color != "green" else "white")
            btn.grid(row=3 + idx // 3, column=idx % 3, pady=10, padx=10)

    def create_treeview(self):
        self.tree = ttk.Treeview(self.root, columns=('Product', 'Quantity', 'Price'), show='headings')
        self.tree.heading('Product', text='Product Name')
        self.tree.heading('Quantity', text='Quantity')
        self.tree.heading('Price', text='Price')
        self.tree.grid(row=6, column=0, columnspan=3, padx=10, pady=10)

    def clear_entries(self):
        for entry in self.entries:
            entry.delete(0, tk.END)

    def get_product_details(self):
        name = self.entries[0].get()
        quantity = self.entries[1].get()
        price = self.entries[2].get()
        return name, quantity, price

    def add_product(self):
        name, quantity, price = self.get_product_details()
        if name and quantity and price:
            self.manager.add_product(name, quantity, price)
            self.clear_entries()
            self.preview_inventory()  # Refresh the table
        else:
            messagebox.showerror("Input Error", "All fields are required.")

    def edit_product(self):
        name, quantity, price = self.get_product_details()
        if name and quantity and price:
            self.manager.edit_product(name, quantity, price)
            self.clear_entries()
            self.preview_inventory()  # Refresh the table
        else:
            messagebox.showerror("Input Error", "All fields are required.")

    def delete_product(self):
        name, *_ = self.get_product_details()
        if name:
            self.manager.delete_product(name)
            self.clear_entries()
            self.preview_inventory()  # Refresh the table
        else:
            messagebox.showerror("Input Error", "Product name is required.")

    def low_stock_alert(self):
        self.manager.low_stock_alert()

    def preview_inventory(self):
        # Clear current treeview contents
        for item in self.tree.get_children():
            self.tree.delete(item)

        # Insert updated inventory data into the treeview
        for product in self.manager.inventory.values():
            self.tree.insert('', 'end', values=(product.name, product.quantity, product.price))

# Main Program
if __name__ == "__main__":
    root = tk.Tk()
    app = InventoryApp(root)
    root.mainloop()